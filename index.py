from flask import Flask, request, render_template
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import time
from tqdm import tqdm #barras de progreso 
#from keras import models 
#from keras import layers
import torch
from torch import nn
from torchsummary import summary
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import json
import requests
import traceback
import boto3
import os
from dotenv import load_dotenv
from sklearn.svm import SVR
from sklearn.linear_model import LinearRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import cross_val_score
from openpyxl import Workbook
from fpdf import FPDF
from sklearn.preprocessing import MinMaxScaler
from keras.models import Sequential
from keras.layers import LSTM, Dense

app = Flask(__name__)


data = pd.read_csv("./Google_Stock_Price_Train.csv", index_col='Date', parse_dates=['Date'])
obj_cols = ['Volume', 'Close'] #convertir variables a numericas
data[obj_cols] = data[obj_cols].replace('[\,,]', '', regex=True).astype(float)

X= np.array(data.iloc[:, [0, 1, 2, 4]])
y= np.array(data['Close'])
        
sc = MinMaxScaler(feature_range=(0,1))
sc_x = MinMaxScaler(feature_range=(0,1))

y = y.reshape(-1, 1)
y= sc.fit_transform(y)

X= sc_x.fit_transform(X)
        
X_train, X_test, Y_train, Y_test = train_test_split(X, y, test_size = 0.2, random_state=40)

X_train = np.reshape(X_train, (X_train.shape[0], X_train.shape[1], 1))
X_test = np.reshape(X_test, (X_test.shape[0], X_test.shape[1], 1))        
        
modelo = Sequential()
modelo.add(LSTM(60,activation = 'relu', return_sequences=True, input_shape=(X_train.shape[1],1)))
modelo.add(LSTM(60,activation = 'relu', return_sequences=True))

modelo.add(LSTM(80))
#modelo.add(LSTM(32))
modelo.add(Dense(40))
modelo.add(Dense(1))
modelo.summary()
modelo.compile(optimizer='adam', loss='mean_squared_error')
modelo.fit(X_train,Y_train, batch_size=10, epochs=80) #20 y 100



class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'DataFrame a PDF', 0, 1, 'C')
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, 'Página %s' % self.page_no(), 0, 0, 'C')
    
    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, title, 0, 1, 'L')
        self.ln(5)
    
    def chapter_body(self, df):
        self.set_font('Arial', '', 10)
        for index, row in df.iterrows():
            for col in df.columns:
                self.cell(40, 10, str(row[col]), 1, 0, 'L')
            self.ln()

@app.route('/', methods=['POST'])
def hello():
    body = request.json
    print(body)
    numWhatsapp = body['message']['owner']
    content = body['message']['content']
    Siguiente_caja1 = body['Siguiente_caja1']
    Siguiente_caja0 = body['Siguiente_caja0']
    print(Siguiente_caja0)
    try:
        data = pd.read_csv('./Google_Stock_Price_Train.csv')
        workbook = Workbook()
        sheet = workbook.active
        for row in data.itertuples():
            sheet.append(row[1:])
        workbook.save('Google_Stock_Price_Train.xlsx')
        dataframe = pd.read_excel('./Google_Stock_Price_Train.xlsx')
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        pdf.cell(0, 10, "Fecha              Apertura              P. Alto              P.Bajo              Cierre              Volumen", ln=True, align="L")
        for row in dataframe.iterrows():
            _, data = row
            for value in data:
                # Agregar cada valor al PDF
                pdf.cell(30, 10, str(value), border=1)
            pdf.ln()
        pdf.output("Google_Stock_Price_Train.pdf")

        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/Google_Stock_Price_Train.pdf'
        s3_client.upload_file('Google_Stock_Price_Train.pdf', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'application/pdf'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)
        data = pd.read_csv('./Google_Stock_Price_Train.csv')

        print("esto es el head")
        print(data.head())
        print("esto es la info")
        print(data.info())
        print("esto es la suma")
        print(data.isnull().sum())

        obj_cols = ['Volume', 'Close'] #convertir variables a numericas
        data[obj_cols] = data[obj_cols].replace('[\,,]', '', regex=True).astype(float)

        print("esto es la suma despues de procesar")
        print(data.isnull().sum())
        #dropping the customerID column
        data=data.drop('Date',axis=1)
        print("borrando la columna customerID")
        print(data.info())

        #se va a predecir la variable CLOSE para los precios de cierre
        x = np.array(data.drop('Close',axis=1))
        y = data['Close']
        print("se va a predecir la variable CLOSE para los precios de cierre")
        print(x.shape,y.shape)

        x = StandardScaler().fit_transform(x)
        print(x)
        print("resultado de la variable x")
        print(x.shape)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "document",
                        "document":{
                            "link": object_url,
                            "filename": "Acciones G."
                        }
                    },
                    "body": {
                        "text": "Al final de todos los procesos, se va a predecir el porcentaje de cierre de las acciones de Google ⌛\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "dispersion",
                                    "title": "Grafico Dispersión"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja1})
        }

        return response
    except Exception as e:
        traceback.print_exc()
        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "document",
                        "document":{
                            "link": "https://exportacion.s3.amazonaws.com/postobon-corporativo/Google_Stock_Price_Train.pdf",
                            "filename": "Acciones G."
                        }
                    },
                    "body": {
                        "text": "Al final de todos los procesos, se va a predecir el porcentaje de cierre de las acciones de Google ⌛\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "dispersion",
                                    "title": "Grafico Dispersión"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja0})
        }
        return response


@app.route('/dispersion', methods=['POST'])
def dispersion():
    body = request.json
    numWhatsapp = body['message']['owner']
    content = body['message']['content']
    Siguiente_caja1 = body['Siguiente_caja1']
    Siguiente_caja0 = body['Siguiente_caja0']
    try:
        data = pd.read_csv('./Google_Stock_Price_Train.csv')
        
        obj_cols = ['Volume', 'Close'] #convertir variables a numericas
        data[obj_cols] = data[obj_cols].replace('[\,,]', '', regex=True).astype(float)

        data=data.drop('Date',axis=1)
        x = np.array(data.drop('Close',axis=1))
        y = data['Close']
        
        x = StandardScaler().fit_transform(x)

        x_train, x_test, y_train, y_test = train_test_split(x,y, test_size=0.2, random_state=40)
        plt.scatter(x_train[:,3], y_train, c='salmon',s=15)
        plt.scatter(x_test[:,3], y_test, c='g',s=10)
        plt.savefig('graph.png')
        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/graph.png'
        s3_client.upload_file('graph.png', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'image/png'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "image",
                        "image":{
                            "link": object_url
                        }
                    },
                    "body": {
                        "text": "Esta imagen representa la Grafica de Dispersión de los datos de las acciones de Google\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "continuar",
                                    "title": "Predecir Acción"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja1})
        }

        return response
    except Exception as e:
        traceback.print_exc()
        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "image",
                        "image":{
                            "link": "https://exportacion.s3.amazonaws.com/postobon-corporativo/graph.png"
                        }
                    },
                    "body": {
                        "text": "Esta imagen representa la Grafica de Dispersión de los datos de las acciones de Google\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "continuar",
                                    "title": "Predecir Acción"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja1})
        }
        return response


@app.route('/barras', methods=['POST'])
def barras():
    body = request.json
    numWhatsapp = body['message']['owner']
    content = body['message']['content']
    Siguiente_caja1 = body['Siguiente_caja1']
    Siguiente_caja0 = body['Siguiente_caja0']
    try:
        data = pd.read_csv('./Google_Stock_Price_Train.csv')
        
        obj_cols = ['Volume', 'Close'] #convertir variables a numericas
        data[obj_cols] = data[obj_cols].replace('[\,,]', '', regex=True).astype(float)

        data=data.drop('Date',axis=1)
        x = np.array(data.drop('Close',axis=1))
        y = data['Close']
        
        x = StandardScaler().fit_transform(x)

        x_train, x_test, y_train, y_test = train_test_split(x,y, test_size=0.2, random_state=40)
        models = [LinearRegression(), DecisionTreeRegressor(), RandomForestRegressor()]

        # Use cross-validation to compute the R-squared score for each model
        cv_scores = []
        for model in models:
            scores = cross_val_score(model, x_train, y_train, cv=5, scoring='r2', n_jobs=-1)
            cv_scores.append(scores.mean())

        # Plot the results
        fig, ax = plt.subplots(figsize=(15, 6))
        rects = ax.bar(['Linear', 'Decision Tree', 'Random Forest'], cv_scores, color='orange')
        ax.set_ylim(0, 1)
        ax.set_title('R2 Comparison of Regression Models')
        ax.set_xlabel('Model')
        ax.set_ylabel('R-squared')

        # Add labels above each bar
        for rect in rects:
            height = rect.get_height()
            ax.text(rect.get_x() + rect.get_width()/2., height, f'{height:.5f}', ha='center', va='bottom')

        # Show the plot
        plt.savefig('barras.png')
        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/barras.png'
        s3_client.upload_file('barras.png', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'image/png'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "image",
                        "image":{
                            "link": object_url
                        }
                    },
                    "body": {
                        "text": "Esta imagen representa la Grafica de Barras de los datos de las acciones de Google\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "continuar",
                                    "title": "Predecir Acción"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja1})
        }

        return response
    except Exception as e:
        traceback.print_exc()
        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "interactive",
                "interactive": {
                    "type": "button",
                    "header":{
                        "type": "image",
                        "image":{
                            "link": "https://exportacion.s3.amazonaws.com/postobon-corporativo/barras.png"
                        }
                    },
                    "body": {
                        "text": "Esta imagen representa la Grafica de Barras de los datos de las acciones de Google\n\nEscoge una de las opciones 👇"
                    },
                    "action": {
                        "buttons": [
                            {
                                "type": "reply",
                                "reply": {
                                    "id": "continuar",
                                    "title": "Predecir Acción"
                                }
                            }               
                        ]
                    }
                }
            })
        )
        print(response.json())
        response = {
            "statusCode": 200,
            "body": json.dumps({"nextStateId": Siguiente_caja1})
        }
        return response

@app.route('/api/endpoint', methods=['GET'])
def get_data():
    # Lógica para obtener los datos
    return 'Datos obtenidos'

@app.route('/final', methods=['POST'])
def create_data():
    body = request.json
    numWhatsapp = body['message']['owner']
    content = body['message']['content']
    Siguiente_caja1 = body['Siguiente_caja1']
    Siguiente_caja0 = body['Siguiente_caja0']
    try:
        
        fig, ax = plt.subplots(1, 1, figsize=(23, 8))
        fig.set_facecolor('#36AE7C')
        ax.plot(data['Close'], marker="$♥$",markersize=5, linewidth=2,linestyle='solid', color="#187498")
        ax.plot(data['Open'], marker = "$♥$",markersize=5,linewidth=2,linestyle='solid' ,color="#EB5353")
        plt.title('Historial de Precios')
        plt.xlabel('_Fecha_')
        plt.ylabel('_Precio_')
        ax.legend(['Cierre','Apertura'])
        ax.set_facecolor("#ffffff")

        plt.savefig('rayas.png')
        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/rayas.png'
        s3_client.upload_file('rayas.png', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'image/png'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "image",
                "image": {
                    "link": object_url,
                    "caption": "Historial precios"
                }
            })
        )
        print(response.json())


        
        predicciones = modelo.predict(X_test)

        rmse = np.sqrt(mean_squared_error(Y_test, predicciones))

        # Calcular R2
        r2 = r2_score(Y_test, predicciones)
        print("R2:", r2)
        print("RMSE:", rmse)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "text",
                "text": {
                    "body": "🔃 Ajuste de los datos a la curva "+ str(r2) +"\nEl error cuadratico medio es: " + str(rmse)
                }
            })
        )
        print(response.json())

        fig, ax = plt.subplots(1, 1, figsize=(23, 8))
        fig.set_facecolor('#36AE7C')
        ax.plot(predicciones, marker = "$♥$",markersize=6,linewidth=2,linestyle='solid' ,color="#EB5353")
        ax.plot(Y_test[:,0], marker = "$♥$",markersize=6,linewidth=2,linestyle='solid' ,color="#000000", alpha=0.4)
        plt.title('Historial de Precios')
        plt.xlabel('_X_')
        plt.ylabel('_Y_')
        ax.legend(['Cierre','predicción'])
        ax.set_facecolor("#ffffff")

        plt.savefig('predic.png')
        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/predic.png'
        s3_client.upload_file('predic.png', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'image/png'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "image",
                "image": {
                    "link": object_url,
                    "caption": "Predicción"
                }
            })
        )
        print(response.json())
        
        datos_nuevos = np.array([[0.09701243, 0.09834351, 0.09823458, 0.22993592]])  # Ingresa los valores de Open, High, Low y Volume
        datos_nuevos2 = np.array([[0.09701243, 0.09834351, 0.09823458, 0.22993592]])
        variables = ['Open', 'High', 'Low', 'Volume']
        vals= []

        for i,col in enumerate(variables):
          val = (datos_nuevos2[0][i] * (data[col].max() - data[col].min())) + data[col].min() #Open, High, Low y Volume
          vals.append(val)



        datos_nuevos = np.reshape(datos_nuevos, (datos_nuevos.shape[0], datos_nuevos.shape[1], 1))

        prediccion = modelo.predict(datos_nuevos)

        close_predicho = prediccion[0, 0]

        min = data['Close'].min()
        rango = data['Close'].max() - data['Close'].min()
        
        close_pred = ((close_predicho * rango) + min)
        
        print(close_pred)

        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({
                "recipient_type": "individual",
                "to": numWhatsapp,
                "type": "text",
                "text": {
                    "body": "💵 A partir de los datos:\n\n*Apertura:* "+str(vals[0])+"\n*P.Alto:* "+str(vals[1])+"\n*P.Bajo:* "+str(vals[2])+"\n*Volumen:* "+str(vals[3])+"\n\nEl valor predicho para el Cierre es de: 💲" + str(close_pred) + "\nEl valor normalizado es de: " + str(close_predicho)+"☑️"
                }
            })
        )
        print(response.json())

        df_results = pd.DataFrame({'Close': sc.inverse_transform(Y_test.reshape(-1, 1)).flatten(),
                           'Predictions': sc.inverse_transform(predicciones).flatten()})

        pdf = PDF()
        pdf.add_page()

        # Definir el título del capítulo y el contenido del DataFrame
        pdf.chapter_title('Precio Cierre          Precio Predicho')
        pdf.chapter_body(df_results)

        # Guardar el archivo PDF
        pdf.output('final.pdf')

        access_key = os.getenv('ACCESS_KEY')
        secret_key = os.getenv('ACCESS_SECRET')
        region = 'us-east-1' 
        s3_client = boto3.client('s3', aws_access_key_id=access_key, aws_secret_access_key=secret_key, region_name=region)
        bucket_name = os.getenv('BUCKET')
        object_key = 'postobon-corporativo/final.pdf'
        s3_client.upload_file('final.pdf', bucket_name, object_key, ExtraArgs={'ACL': 'public-read', 'ContentType': 'application/pdf'})
        bucket_url = f'https://{bucket_name}.s3.amazonaws.com/'
        object_url = bucket_url + object_key

        print('URL del objeto subido:', object_url)
        
        response = requests.post("https://waba.360dialog.io/v1/messages", headers={
                    "D360-Api-Key": os.getenv('API_KEY'),
                    "Content-Type": "application/json",
                }
                , data=json.dumps({ 
                    "to": numWhatsapp,
	                "type": "document",
	                "document": {
	                	"link": object_url,
	                	"filename": "Resultado"
	                }
                })
        )
        print(response.json())

        response = {
                "statusCode": 200,
                "body": json.dumps({"nextStateId": Siguiente_caja1})
        }
        return response
    except Exception as e:
        traceback.print_exc()
        response = {
                "statusCode": 200,
                "body": json.dumps({"nextStateId": Siguiente_caja1})
        }
        return response

if __name__ == '__main__':
    app.run()